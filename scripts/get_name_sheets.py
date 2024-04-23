import sys
import json
import traceback
from openpyxl import load_workbook

def get_name_sheets(excel_path: str) -> list[str]:
    wb = load_workbook(excel_path)
    return wb.sheetnames


if __name__ == "__main__":
    excel_path = sys.argv[1]

    try:
        name_sheets = get_name_sheets(excel_path)
        print(json.dumps(name_sheets))
    except:
        print(traceback.format_exc())
